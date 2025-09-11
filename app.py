import os
import tempfile
from flask import Flask, render_template, request, send_file, jsonify
import pdfplumber
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# --- CONFIGURATION ---
# The logo must be in the same folder as this python script.
logo_file_path = os.path.join(os.path.dirname(__file__), 'logo.png')


def draw_label_in_excel(sheet, start_row, start_col, data):
    """
    This function draws a single, formatted label.
    This is the styling from your final desktop script.
    """
    bold_font_large = Font(name='Arial', size=12, bold=True) [cite: 1]
    bold_font_qty = Font(name='Arial', size=12, bold=True) [cite: 1]
    plain_font_medium = Font(name='Arial', size=9, bold=False) [cite: 1]
    plain_font_small = Font(name='Arial', size=10, bold=False) [cite: 1]
    header_font = Font(name='Arial', size=11, bold=True) [cite: 1]
    arrow_font = Font(name='Arial', size=48, bold=True) [cite: 1, 2]
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True) [cite: 2]
    align_left_no_wrap = Alignment(horizontal='left', vertical='center', wrap_text=False) [cite: 2]

    # Set row heights for this specific label
    heights = [36, 35, 22, 25, 19, 33] [cite: 2]
    for i in range(6):
        sheet.row_dimensions[start_row + i].height = heights[i] [cite: 2]

    # Merge cells relative to the starting column
    col_offset = start_col - 1
    sheet.merge_cells(start_row=start_row, start_column=1 + col_offset, end_row=start_row, end_column=1 + col_offset) [cite: 2]
    sheet.merge_cells(start_row=start_row, start_column=2 + col_offset, end_row=start_row, end_column=4 + col_offset) [cite: 2]
    sheet.merge_cells(start_row=start_row + 1, start_column=1 + col_offset, end_row=start_row + 1, end_column=1 + col_offset) [cite: 2]
    sheet.merge_cells(start_row=start_row + 1, start_column=2 + col_offset, end_row=start_row + 1, end_column=3 + col_offset) [cite: 2, 3]
    sheet.merge_cells(start_row=start_row + 2, start_column=1 + col_offset, end_row=start_row + 2, end_column=1 + col_offset) [cite: 3]
    sheet.merge_cells(start_row=start_row + 2, start_column=2 + col_offset, end_row=start_row + 2, end_column=3 + col_offset) [cite: 3]
    sheet.merge_cells(start_row=start_row + 3, start_column=1 + col_offset, end_row=start_row + 3, end_column=1 + col_offset) [cite: 3]
    sheet.merge_cells(start_row=start_row + 3, start_column=2 + col_offset, end_row=start_row + 3, end_column=3 + col_offset) [cite: 3]
    sheet.merge_cells(start_row=start_row + 4, start_column=1 + col_offset, end_row=start_row + 5, end_column=1 + col_offset) [cite: 3]
    sheet.merge_cells(start_row=start_row + 4, start_column=2 + col_offset, end_row=start_row + 5, end_column=2 + col_offset) [cite: 3]
    sheet.merge_cells(start_row=start_row + 4, start_column=3 + col_offset, end_row=start_row + 5, end_column=3 + col_offset) [cite: 4]

    # Place data and apply styles
    cell = sheet.cell(row=start_row, column=2 + col_offset, value="JOONHEE ENGINEERING SDN. BHD.") [cite: 4, 5]
    cell.font = header_font [cite: 5]
    cell.alignment = align_left_no_wrap [cite: 5]
    cell = sheet.cell(row=start_row + 1, column=1 + col_offset, value="JOONHEE ENGINEERING") [cite: 5]
    cell.font = plain_font_medium [cite: 5]
    cell.alignment = align_center [cite: 5]
    cell = sheet.cell(row=start_row + 1, column=2 + col_offset, value="→") [cite: 5]
    cell.font = arrow_font [cite: 5]
    cell.alignment = align_center [cite: 5]
    qty_text = f"AISB\nQTY: {data['qty']}" [cite: 5]
    qty_cell = sheet.cell(row=start_row + 1, column=4 + col_offset, value=qty_text) [cite: 5]
    qty_cell.font = bold_font_qty [cite: 5]
    qty_cell.alignment = align_center [cite: 5]
    sheet.cell(row=start_row + 2, column=1 + col_offset, value="PART NO.").font = plain_font_small [cite: 5, 6]
    sheet.cell(row=start_row + 2, column=2 + col_offset, value="KANBAN NO.").font = plain_font_small [cite: 6]
    sheet.cell(row=start_row + 2, column=4 + col_offset, value="ISSUE DATE :").font = plain_font_small [cite: 6]
    sheet.cell(row=start_row + 4, column=1 + col_offset, value="PART NAME").font = plain_font_small [cite: 6]
    sheet.cell(row=start_row + 4, column=2 + col_offset, value="COLOR CODE").font = plain_font_small [cite: 6]
    sheet.cell(row=start_row + 4, column=4 + col_offset, value="DELIVERY DATE :").font = plain_font_small [cite: 6]
    sheet.cell(row=start_row + 3, column=1 + col_offset, value=data['part_no']).font = bold_font_large [cite: 6]
    sheet.cell(row=start_row + 3, column=2 + col_offset, value=data['kanban_no']).font = bold_font_large [cite: 6]
    sheet.cell(row=start_row + 3, column=4 + col_offset, value=data['issue_date']).font = bold_font_large [cite: 7]
    sheet.cell(row=start_row + 5, column=4 + col_offset, value=data['delivery_date']).font = bold_font_large [cite: 7]
    sheet.cell(row=start_row + 4, column=1 + col_offset, value=data['part_name']).font = plain_font_medium [cite: 7]

    for r in range(start_row + 2, start_row + 6):
        for c in range(1 + col_offset, 5 + col_offset):
            sheet.cell(row=r, column=c).alignment = align_center [cite: 7]
            
    # Apply borders using the granular logic from the desktop script
    medium_side = Side(style='medium') [cite: 7]
    thin_side = Side(style='thin') [cite: 7]
    for r_offset in range(6):
        for c_offset in range(4): [cite: 8]
            r, c = start_row + r_offset, start_col + c_offset [cite: 8]
            current_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side) [cite: 8]
            if r_offset == 0: current_border.top = medium_side [cite: 8]
            if r_offset == 5: current_border.bottom = medium_side [cite: 8]
            if c_offset == 0: current_border.left = medium_side [cite: 8]
            if c_offset == 3: current_border.right = medium_side [cite: 9]
            if r_offset == 0: current_border.bottom = medium_side [cite: 9]
            sheet.cell(row=r, column=c).border = current_border [cite: 9]
            
    # Insert Logo
    if os.path.exists(logo_file_path):
        img = Image(logo_file_path) [cite: 9]
        img.height = 45 [cite: 9]
        img.width = 75 [cite: 9]
        logo_anchor = f'{get_column_letter(start_col)}{start_row}' [cite: 9]
        sheet.add_image(img, logo_anchor) [cite: 10]
    else:
        sheet.cell(row=start_row, column=1 + col_offset, value="Logo not found").alignment = align_center [cite: 10]

def setup_sheet_for_printing(sheet):
    """Applies all page setup options to a single sheet."""
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4 [cite: 10]
    sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT [cite: 10]
    sheet.page_margins.left = 0.2 [cite: 10]
    sheet.page_margins.right = 0.2 [cite: 10]
    sheet.page_margins.top = 0.2 [cite: 10]
    sheet.page_margins.bottom = 0.2 [cite: 10]
    sheet.page_margins.header = 0 [cite: 10]
    sheet.page_margins.footer = 0 [cite: 10]
    sheet.page_setup.scale = 92 [cite: 10]
    
    # Set column widths for the entire two-column grid layout
    sheet.column_dimensions['A'].width = 17 [cite: 11]
    sheet.column_dimensions['B'].width = 8 [cite: 11]
    sheet.column_dimensions['C'].width = 10 [cite: 11]
    sheet.column_dimensions['D'].width = 17 [cite: 11]
    sheet.column_dimensions['E'].width = 2 # Gap column [cite: 11]
    sheet.column_dimensions['F'].width = 17 [cite: 11]
    sheet.column_dimensions['G'].width = 8 [cite: 11]
    sheet.column_dimensions['H'].width = 10 [cite: 11]
    sheet.column_dimensions['I'].width = 17 [cite: 11]

def process_pdfs(files):
    # 1. --- Extract all label data from PDFs first using the accurate desktop method ---
    all_labels_to_print = []
    for pdf_file in files:
        print(f"Reading data from: {pdf_file.filename}")
        try:
            with pdfplumber.open(pdf_file) as pdf:
                all_item_lines_with_page = []
                for page_num, page in enumerate(pdf.pages):
                    text = page.extract_text() or "" [cite: 12]
                    for line in text.split('\n'):
                        if re.match(r'^\d+\s+[A-Z0-9\-]+\s+.*EA$', line.strip()): [cite: 13]
                            all_item_lines_with_page.append({'line': line.strip(), 'page_num': page_num}) [cite: 13]
                if not all_item_lines_with_page:
                    print(f"      WARNING: No items found in {pdf_file.filename}") [cite: 13]
                    continue [cite: 14]
                for i, item_info in enumerate(all_item_lines_with_page):
                    line = item_info['line'] [cite: 14]
                    page_num = item_info['page_num'] [cite: 14]
                    page_text = pdf.pages[page_num].extract_text() or "" [cite: 14]
                    issue_date_match = re.search(r"Issue Date\s*:\s*(\d{2}-\w{3}-\d{4})", page_text) [cite: 14]
                    delivery_date_match = re.search(r"Delivery Date\s*:\s*(\d{2}-\w{3}-\d{4})", page_text) [cite: 15]
                    current_issue_date = issue_date_match.group(1) if issue_date_match else "N/A" [cite: 15]
                    current_delivery_date = delivery_date_match.group(1) if delivery_date_match else "N/A" [cite: 15]
                    parts = line.split() [cite: 15]
                    std_pkg_qty = parts[-3] [cite: 15]
                    description_tokens = parts[2:-3] [cite: 16]
                    if len(parts) > 2 and re.match(r'[A-Z]-\d{3,}', parts[2]):
                        part_no = parts[2] [cite: 16]
                    else:
                        internal_code = parts[1] [cite: 16]
                        match = re.search(r'([A-Z])[A-Z0-9]*-?([0-9]{3,})', internal_code) [cite: 17]
                        part_no = f"{match.group(1)}-{match.group(2)}" if match else internal_code [cite: 17]
                    if description_tokens and description_tokens[0] == part_no:
                        description_tokens = description_tokens[1:] [cite: 17]
                    description = " ".join(description_tokens) [cite: 18]
                    page_lines = [l.strip() for l in page_text.split('\n') if l.strip()] [cite: 18]
                    start_idx_on_page = page_lines.index(line) + 1 [cite: 18]
                    end_idx_on_page = len(page_lines) [cite: 18]
                    for next_item_info in all_item_lines_with_page[i + 1:]:
                        if next_item_info['page_num'] == page_num and next_item_info['line'] in page_lines: [cite: 19]
                            end_idx_on_page = page_lines.index(next_item_info['line']) [cite: 19]
                            break [cite: 19]
                    block_lines = page_lines[start_idx_on_page:end_idx_on_page] [cite: 19]
                    kanban_cards = [] [cite: 20]
                    for bl in block_lines:
                        kanban_cards.extend(re.findall(r'\b\d{10}\b', bl)) [cite: 20]
                    if not kanban_cards:
                        print(f"      WARNING: No Kanban cards found for part {part_no}") [cite: 20]
                        continue [cite: 21]
                    for card in kanban_cards:
                        label_info = {
                            "part_no": part_no, "part_name": description, "qty": std_pkg_qty,
                            "kanban_no": card, "issue_date": current_issue_date, "delivery_date": current_delivery_date
                        } [cite: 21, 22]
                        all_labels_to_print.append(label_info) [cite: 22]
        except Exception as e:
            print(f"ERROR while processing {pdf_file.filename}: {e}") [cite: 22]

    # 2. --- Create the Excel file and generate labels using the multi-sheet method ---
    print(f"\nFound {len(all_labels_to_print)} total labels. Generating Excel file...")
    workbook = openpyxl.Workbook() [cite: 23]
    workbook.remove(workbook.active) # Remove the default sheet 

    # Process labels in chunks of 10 (one page per sheet)
    for page_index in range(0, len(all_labels_to_print), 10):
        page_labels = all_labels_to_print[page_index : page_index + 10] [cite: 23]
        
        # Create a new sheet for this page
        sheet_name = f"Page {page_index // 10 + 1}" [cite: 23]
        sheet = workbook.create_sheet(title=sheet_name) [cite: 23]
        print(f"Creating sheet: {sheet_name}") [cite: 23]
        
        # Setup the page layout for this new sheet
        setup_sheet_for_printing(sheet) [cite: 23]
        
        current_print_row = 1 [cite: 23]
        # Draw the 10 labels for this page in a 2x5 grid
        for i, label_data in enumerate(page_labels):
            start_col = 1 if i % 2 == 0 else 6 [cite: 24]
            
            draw_label_in_excel(sheet, current_print_row, start_col, label_data) [cite: 24]
            
            # If the right-side label is done, move to the next row of labels
            if i % 2 == 1:
                current_print_row += 6 # Move past the 6 rows of the label [cite: 25]
                # Add a gap row, unless it's the very last label on the page
                if i < len(page_labels) - 1:
                    sheet.row_dimensions[current_print_row].height = 12 [cite: 25]
                    current_print_row += 1 [cite: 26]

    # 3. --- Save and return the file ---
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    workbook.save(temp_file.name)
    temp_file.close()
    return temp_file.name


@app.route('/')
def home():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_and_process():
    if 'files' not in request.files:
        return jsonify({"success": False, "message": "No files uploaded"})

    files = request.files.getlist('files')
    pdf_files = [file for file in files if file.filename.lower().endswith('.pdf')] [cite: 46, 47]

    if not pdf_files:
        return jsonify({"success": False, "message": "No valid PDF files"}) [cite: 47]

    excel_path = process_pdfs(pdf_files)
    
    # Create a dynamic filename with a timestamp
    current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_filename = f"Generated_Labels_{current_time}.xlsx"
    
    return send_file(excel_path, as_attachment=True, download_name=output_filename)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
